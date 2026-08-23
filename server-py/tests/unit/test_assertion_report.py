"""Stage F4: 断言报告导出端点（POST .../assertion-report）与 build_assertion_report。

覆盖（对应 implement.md F4 gate）：
- JSON / XLSX 两种格式生成（201 + artifact 落库 + 文件内容校验）；
- actual 经 redact_run_value 脱敏（含 secret 明文不落盘）；
- 失败截图 / trace 引用装配：按 `failure-step-{序号}.png` 匹配到 stepIndex，
  缺失的截图留空不报错；trace.zip 全量引用；
- 状态码：201 / 400（非法格式）/ 403（无工作区访问权）/ 404（run 不存在）/
  409（run 无断言）。
"""

from __future__ import annotations

import asyncio
import json
from pathlib import Path

import pytest
from fastapi import Request
from openpyxl import load_workbook

from autoflow.core import now
from autoflow.handler import create_platform_router
from autoflow.http import PlatformError
from autoflow.services import AuthUser, PlatformServices

PROJECT_ID = "proj-report"
RUN_ID = "run-report-1"
REPORT_ROUTE = (
    "/api/platform/projects/{project_id}/runs/{run_id}/assertion-report"
)

ASSERTIONS = [
    {
        "stepIndex": 0,
        "stepId": "s1",
        "title": "可见性断言",
        "type": "visibility",
        "passed": True,
        "expected": "visible",
        "actual": "visible",
        "durationMs": 10,
    },
    {
        "stepIndex": 1,
        "stepId": "s2",
        "title": "数量断言",
        "type": "count",
        "passed": False,
        "expected": "3",
        "actual": "2",
        "durationMs": 12,
    },
]

SNAPSHOT = {
    "flow": {"id": "flow-report", "name": "报告流程", "steps": []},
    "environment": {"id": "env-report", "name": "报告环境"},
}


def _seed_services(tmp_path) -> tuple[PlatformServices, dict]:
    services = PlatformServices(str(tmp_path))
    admin = services.bootstrap_super_admin(
        "report-admin@example.test", "Report Admin", "report-password"
    )
    session = services.create_auth_session(admin)
    workspace = services.create_workspace(admin, "Report workspace")
    services.database.execute(
        """
        INSERT INTO platform_projects (
          id, workspace_id, slug, name, description, created_at, updated_at
        ) VALUES (?, ?, ?, ?, ?, ?, ?)
        """,
        (
            PROJECT_ID,
            workspace["id"],
            PROJECT_ID,
            "Report project",
            "",
            now(),
            now(),
        ),
    )
    services.database.execute(
        """
        INSERT INTO agents (
          id, workspace_id, name, credential_hash, status, browser_version,
          os, max_concurrency, created_at
        ) VALUES ('agent-report', ?, 'Report agent', 'hash', 'online', 'Chromium',
                  'linux', 1, ?)
        """,
        (workspace["id"], now()),
    )
    services.database.execute(
        """
        INSERT INTO flow_revisions (
          id, project_id, revision_number, status, flow_snapshot,
          environment_snapshot, element_snapshot, dataset_snapshot,
          checksum, created_by, created_at
        ) VALUES ('rev-report', ?, 1, 'published', '{}', '{}', '[]',
                  '{}', 'checksum', 'user-report', ?)
        """,
        (PROJECT_ID, now()),
    )
    return services, session


def _seed_run(
    services: PlatformServices,
    *,
    run_id: str = RUN_ID,
    assertions: list[dict] | None = ASSERTIONS,
    status: str = "success",
    actual_overrides: dict[int, str] | None = None,
) -> None:
    if actual_overrides:
        assertions = [
            dict(item, actual=actual_overrides.get(idx, item["actual"]))
            for idx, item in enumerate(assertions or [])
        ]
    result = {"status": status}
    if assertions is not None:
        result["assertions"] = assertions
    services.database.execute(
        """
        INSERT INTO platform_runs (
          id, project_id, revision_id, environment_id, agent_id, executor_type,
          status, snapshot, result, created_by, created_at, updated_at
        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """,
        (
            run_id,
            PROJECT_ID,
            "rev-report",
            "env-report",
            "agent-report",
            "managed",
            status,
            json.dumps(SNAPSHOT, ensure_ascii=False),
            json.dumps(result, ensure_ascii=False),
            "user-report",
            now(),
            now(),
        ),
    )


def _seed_artifact(
    services: PlatformServices, artifact_id: str, name: str, content_type: str
) -> None:
    artifact_directory = services.managed_runner.artifact_directory
    artifact_directory.mkdir(parents=True, exist_ok=True)
    path = artifact_directory / f"asset-{artifact_id}"
    path.write_bytes(b"\x00asset")
    services.database.execute(
        """
        INSERT INTO platform_artifacts (
          id, run_id, project_id, name, content_type, path, created_at
        ) VALUES (?, ?, ?, ?, ?, ?, ?)
        """,
        (artifact_id, RUN_ID, PROJECT_ID, name, content_type, str(path), now()),
    )


def _route(services: PlatformServices):
    router = create_platform_router(services)
    return next(
        route for route in router.routes if getattr(route, "path", None) == REPORT_ROUTE
    )


def _call(route, *, token: str, query_string: bytes = b"", **path_params):
    async def run():
        async def receive():
            return {"type": "http.request", "body": b"", "more_body": False}

        scope = {
            "type": "http",
            "asgi": {"version": "3.0"},
            "http_version": "1.1",
            "method": "POST",
            "scheme": "http",
            "path": "/api",
            "raw_path": b"/api",
            "query_string": query_string,
            "headers": [(b"authorization", f"Bearer {token}".encode())],
            "client": ("127.0.0.1", 1234),
            "server": ("127.0.0.1", 8787),
        }
        return await route.endpoint(Request(scope, receive=receive), **path_params)

    return asyncio.run(run())


def _report_from_artifact(services: PlatformServices, artifact_id: str) -> dict:
    row = services.database.execute(
        "SELECT path FROM platform_artifacts WHERE id = ?", (artifact_id,)
    ).fetchone()
    assert row is not None
    return json.loads(Path(row[0]).read_text(encoding="utf-8"))


def test_assertion_report_json_assembles_assets(tmp_path):
    """JSON 报告：断言行装配 + failure-step-{n}.png 截图/trace 引用（缺失留空）。"""
    services, session = _seed_services(tmp_path)
    try:
        _seed_run(services)
        # 步骤2（index 1）失败截图 + trace；步骤1（index 0）无截图 → 留空。
        _seed_artifact(services, "artifact-shot-2", "failure-step-2.png", "image/png")
        _seed_artifact(services, "artifact-trace", "trace.zip", "application/zip")

        response = _call(
            _route(services),
            token=session["token"],
            query_string=b"format=json",
            project_id=PROJECT_ID,
            run_id=RUN_ID,
        )
        assert response.status_code == 201
        body = json.loads(response.body)
        artifact = body["artifact"]
        assert artifact["contentType"] == "application/json"
        assert artifact["name"].startswith(f"assertion-report-{RUN_ID}.")
        assert artifact["name"].endswith(".json")

        report = _report_from_artifact(services, artifact["id"])
        assert report["runId"] == RUN_ID
        assert report["flowName"] == "报告流程"
        assert report["environmentName"] == "报告环境"
        assert report["status"] == "success"
        assert report["assertionCount"] == 2
        assert report["assertions"] == [
            {
                "stepIndex": 0,
                "stepId": "s1",
                "title": "可见性断言",
                "type": "visibility",
                "passed": True,
                "expected": "visible",
                "actual": "visible",
                "durationMs": 10,
                "screenshotArtifactId": None,
                "traceArtifactId": "artifact-trace",
            },
            {
                "stepIndex": 1,
                "stepId": "s2",
                "title": "数量断言",
                "type": "count",
                "passed": False,
                "expected": "3",
                "actual": "2",
                "durationMs": 12,
                "screenshotArtifactId": "artifact-shot-2",
                "traceArtifactId": "artifact-trace",
            },
        ]
    finally:
        services.close()


def test_assertion_report_xlsx_layout(tmp_path):
    """XLSX 报告：openpyxl 可读，表头 + 逐行数据（序号/判定/引用列）。"""
    services, session = _seed_services(tmp_path)
    try:
        _seed_run(services)
        _seed_artifact(services, "artifact-shot-2", "failure-step-2.png", "image/png")

        response = _call(
            _route(services),
            token=session["token"],
            query_string=b"format=xlsx",
            project_id=PROJECT_ID,
            run_id=RUN_ID,
        )
        assert response.status_code == 201
        body = json.loads(response.body)
        assert body["artifact"]["contentType"] == (
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        assert body["artifact"]["name"].endswith(".xlsx")

        row = services.database.execute(
            "SELECT path FROM platform_artifacts WHERE id = ?",
            (body["artifact"]["id"],),
        ).fetchone()
        workbook = load_workbook(Path(row[0]))
        sheet = workbook.active
        assert sheet.title == "断言报告"
        header = [cell.value for cell in sheet[1]]
        assert header == [
            "序号", "步骤", "类型", "判定", "期望", "实际",
            "耗时(ms)", "失败截图", "Trace",
        ]
        data_rows = list(sheet.iter_rows(min_row=2, values_only=True))
        assert data_rows[0][:4] == (1, "可见性断言", "visibility", "通过")
        assert data_rows[0][4:7] == ("visible", "visible", 10)
        assert data_rows[0][8] in (None, "")  # 无 trace → 留空
        assert data_rows[1][:4] == (2, "数量断言", "count", "失败")
        assert data_rows[1][7] == "artifact-shot-2"
    finally:
        services.close()


def test_assertion_report_redacts_secret_plaintext(tmp_path):
    """actual 含 secret 明文 → 报告内脱敏为 ***，不落明文。"""
    services, session = _seed_services(tmp_path)
    try:
        encrypted = services.encrypt("s3cret-token-42")
        services.database.execute(
            """
            INSERT INTO project_secrets (
              id, project_id, name, key_version, iv, tag, ciphertext,
              created_at, updated_at
            ) VALUES (?, ?, ?, 1, ?, ?, ?, ?, ?)
            """,
            (
                "secret-report",
                PROJECT_ID,
                "API_TOKEN",
                encrypted["iv"],
                encrypted["tag"],
                encrypted["ciphertext"],
                now(),
                now(),
            ),
        )
        _seed_run(services, actual_overrides={0: "token=s3cret-token-42"})

        response = _call(
            _route(services),
            token=session["token"],
            query_string=b"format=json",
            project_id=PROJECT_ID,
            run_id=RUN_ID,
        )
        assert response.status_code == 201
        report = _report_from_artifact(services, json.loads(response.body)["artifact"]["id"])
        assert report["assertions"][0]["actual"] == "token=***"
        assert "s3cret-token-42" not in json.dumps(report, ensure_ascii=False)
    finally:
        services.close()


def test_assertion_report_status_codes(tmp_path):
    """400/403/404/409 状态码覆盖。"""
    services, session = _seed_services(tmp_path)
    try:
        route = _route(services)

        # 409：run 存在但 result 无断言。
        _seed_run(services, assertions=None, status="success")
        with pytest.raises(PlatformError) as error:
            _call(
                route,
                token=session["token"],
                query_string=b"format=json",
                project_id=PROJECT_ID,
                run_id=RUN_ID,
            )
        assert error.value.status == 409
        assert error.value.code == "RUN_HAS_NO_ASSERTIONS"

        # 400：非法 format（复用同一 run，验证 handler 在服务层前拦截）。
        with pytest.raises(PlatformError) as error:
            _call(
                route,
                token=session["token"],
                query_string=b"format=pdf",
                project_id=PROJECT_ID,
                run_id=RUN_ID,
            )
        assert error.value.status == 400
        assert error.value.code == "REPORT_FORMAT_INVALID"

        # 404：run 不存在。
        with pytest.raises(PlatformError) as error:
            _call(
                route,
                token=session["token"],
                query_string=b"format=json",
                project_id=PROJECT_ID,
                run_id="run-missing",
            )
        assert error.value.status == 404
        assert error.value.code == "RUN_NOT_FOUND"

        # 403：无工作区访问权的普通用户。
        stranger = AuthUser("stranger-1", "stranger@example.test", "Stranger")
        services.database.execute(
            """
            INSERT INTO platform_users (id, email, name, created_at)
            VALUES (?, ?, ?, ?)
            """,
            (stranger.id, stranger.email, stranger.name, now()),
        )
        stranger_session = services.create_auth_session(stranger)
        with pytest.raises(PlatformError) as error:
            _call(
                route,
                token=stranger_session["token"],
                query_string=b"format=json",
                project_id=PROJECT_ID,
                run_id=RUN_ID,
            )
        assert error.value.status == 403
    finally:
        services.close()
